import instaloader
import openpyxl


L = instaloader.Instaloader() 
L.login("s1100611_", "yzu1100611")       # login


profile = instaloader.Profile.from_username(L.context, "s1100611_")
followee_list = profile.get_followees() #取得追蹤的帳號清單


wb = openpyxl.Workbook()

wb.create_sheet("工作表1")       # 開啟工作表 1
s1 = wb['工作表1'] 
y=0

for followee in followee_list:
    profile = instaloader.Profile.from_username(L.context, followee.username)
    followers=profile.followers
    print(followee.username, followers)
    row = 2 + y      # 寫入資料的範圍從 row=2 開始 
    s1.cell(row, 1).value = followee.username
    row = 2 + y      # 寫入資料的範圍從 row=2 開始
    s1.cell(row, 2).value = followers
    y+=1
    
        
wb.save('粉絲數高攝影師名單.xlsx')